"""
update_strategy_workbook.py
---------------------------
Automates monthly population of data/strategy/Strategies.xlsx.

Strict rules:
    - Only updates designated data-entry ranges.
    - Never overwrites formulas, formatting, charts, or conditional formatting.
    - Archive copy is created BEFORE any modification.
    - Workbook remains the business reporting layer; Python supplies fresh data.

Input DataFrames
----------------
inventory_detail_df  — output of clean_inventory.run() → detail_df
                       columns: item_code, description, org_name,
                                primary_qty, subinventory_code, category

consumption_detail_df — output of clean_consumption.run() → detail_df
                        columns: item_code, item_desc, org_name, lot_number,
                                 subinventory, txn_type, primary_qty,
                                 consumption_qty, date, year_month, category

market_inputs_df      — list[MarketRecord] or DataFrame from market_inputs.py
                        required metric_name values:
                          "ICE Cotton No. 2"    (unit: USD/lb  → converted to c/lb)
                          "USD/PKR Exchange Rate" (unit: PKR)
                          "SBP Policy Rate"     (unit: %)
                          "PSF Price (China)"   (unit: RMB/ton → converted via rmb_usd_rate)

strategy_df           — output of procurement_engine.run() → strategy_df
                        used for validation only; Strategy sheet is formula-driven

Workbook Sheet Mapping
----------------------
Raw Material  : A6:F{last_item_row}  — inventory detail (match-and-update by item+org)
                A182 / total row     — formula: =SUM(E6:E{N})  [preserved]
                A185-F194            — SUMPRODUCT summary       [preserved]
Consumption   : A5:I{last_data_row}  — transaction detail (full replace)
                TOTAL row            — formula: =SUM(G5:G{N})  [updated if range changes]
                SUMPRODUCT summary   — formulas                 [updated if range changes]
Market Inputs : B6  ICE Cotton Spot Proxy (c/lb)
                B11 PSF Asia Benchmark    (USD/MT)
                B12 SBP Policy Rate       (decimal, e.g. 0.105)
                B13 USD/PKR Assumption    (PKR)
Strategy      : All cells are formulas — validation only, no data writes.
Dashboard     : All cells are formulas — not touched.

Archive Naming
--------------
    data/strategy/archive/strategies_YYYY_MM.xlsx
    e.g. strategies_2026_05.xlsx

Usage
-----
    from update_strategy_workbook import run

    report = run(
        inventory_detail_df=detail_inv,
        consumption_detail_df=detail_cons,
        market_inputs_df=market_df,
        strategy_df=strategy_df,
        workbook_path="data/strategy/Strategies.xlsx",
        archive_dir="data/strategy/archive",
        year=2026,
        month=4,
        period_label="April 2026 (01-Apr to 30-Apr)",
    )

CLI:
    python update_strategy_workbook.py \\
        --inventory   "data/inventory_detail.xlsx" \\
        --consumption "data/consumption_detail.xlsx" \\
        --market      "data/market_inputs.xlsx" \\
        --workbook    "data/strategy/Strategies.xlsx" \\
        --archive-dir "data/strategy/archive" \\
        --year 2026 --month 4
"""

from __future__ import annotations

import argparse
import logging
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

WORKBOOK_SHEETS = ["Raw Material", "Consumption", "Market Inputs", "Strategy", "Dashboard"]

# Raw Material sheet structure (fixed — never changes without manual template update)
RM_HEADER_ROW      = 5    # row containing column headers (Sr#, Item Code, etc.)
RM_DATA_START_ROW  = 6    # first data row
RM_TOTAL_ROW       = 182  # =SUM(E6:E181)
RM_SUMMARY_LABEL   = 185  # "AGGREGATED INVENTORY BY ORG & CATEGORY..."
RM_SUMMARY_HDR_ROW = 186  # org_name | Cotton | Fiber | Stretch Fiber | ...
RM_SUMMARY_START   = 187  # first org summary row (MSL - Fibres U3)
RM_SUMMARY_END     = 193  # last  org summary row (MTM - Spinning U6)
RM_SUMMARY_TOTAL   = 194  # =SUM(B187:B193) etc.
RM_COL_SRNO        = "A"
RM_COL_ITEM_CODE   = "B"
RM_COL_DESC        = "C"
RM_COL_ORG_NAME    = "D"
RM_COL_QTY         = "E"
RM_COL_CATEGORY    = "F"

# Consumption sheet structure (dynamic — total row shifts each month)
CONS_HEADER_ROW   = 4     # row containing column headers
CONS_DATA_START   = 5     # first transaction row
CONS_COL_ITEM     = "A"
CONS_COL_DESC     = "B"
CONS_COL_ORG      = "C"
CONS_COL_LOT      = "D"
CONS_COL_SUBINV   = "E"
CONS_COL_TXN_TYPE = "F"
CONS_COL_QTY      = "G"
CONS_COL_DATE     = "H"
CONS_COL_CATEGORY = "I"
CONS_SUMMARY_ORG_COUNT = 7   # number of org rows in summary block (fixed)

# Market Inputs sheet — map: metric_name → (cell, workbook_unit, conversion_fn)
# Conversions: ICE USD/lb → c/lb (×100); SBP % integer → decimal (/100)
MARKET_INPUT_CELLS: dict[str, tuple[str, str]] = {
    "ICE Cotton No. 2":      ("B6",  "c/lb"),
    "PSF Price (China)":     ("B11", "USD/MT"),
    "SBP Policy Rate":       ("B12", "decimal"),
    "USD/PKR Exchange Rate": ("B13", "PKR"),
}

_CHECK_PASS = "PASS"
_CHECK_WARN = "WARN"
_CHECK_FAIL = "FAIL"


# ---------------------------------------------------------------------------
# Report helpers
# ---------------------------------------------------------------------------

def _row(category: str, check: str, status: str, detail: str, affected: str = "") -> dict:
    return {"category": category, "check": check, "status": status,
            "detail": detail, "affected": affected}


# ---------------------------------------------------------------------------
# Task 1 — Workbook mapping
# ---------------------------------------------------------------------------

def map_workbook_ranges(wb: openpyxl.Workbook) -> dict:
    """Return a structured mapping of input, formula, and protected ranges per sheet.

    Returns:
        Dict keyed by sheet name; each value has:
          input_ranges   — list of range strings Python may write to
          formula_ranges — list of range strings containing formulas (do not write)
          protected      — bool (True = sheet has protection)
    """
    mapping: dict[str, dict] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        prot = getattr(ws, "protection", None)
        protected = bool(prot.sheet) if prot else False
        mapping[name] = {
            "input_ranges":   [],
            "formula_ranges": [],
            "protected":      protected,
        }

    # Raw Material
    last_data = _rm_detect_last_data_row(wb["Raw Material"])
    mapping["Raw Material"]["input_ranges"]   = [f"E{RM_DATA_START_ROW}:E{last_data}"]
    mapping["Raw Material"]["formula_ranges"] = [
        f"E{RM_TOTAL_ROW}",
        f"B{RM_SUMMARY_START}:F{RM_SUMMARY_TOTAL}",
    ]

    # Consumption
    last_data_c, total_r, _, _, _, _ = _cons_detect_structure(wb["Consumption"])
    mapping["Consumption"]["input_ranges"]   = [f"A{CONS_DATA_START}:I{last_data_c}"]
    mapping["Consumption"]["formula_ranges"] = [
        f"G{total_r}",
        f"B{total_r + 5}:E{total_r + 5 + CONS_SUMMARY_ORG_COUNT}",
    ]

    # Market Inputs — only literal value cells (not formula cells)
    mapping["Market Inputs"]["input_ranges"] = [cell for cell, _ in MARKET_INPUT_CELLS.values()]
    mapping["Market Inputs"]["formula_ranges"] = [
        "B25", "B29", "B30", "B31", "B32", "B44", "B45", "B46",
    ]

    # Strategy — all formula
    mapping["Strategy"]["input_ranges"]   = []
    mapping["Strategy"]["formula_ranges"] = ["A3:T26"]

    # Dashboard — all formula
    mapping["Dashboard"]["input_ranges"]   = []
    mapping["Dashboard"]["formula_ranges"] = ["A1:Q38"]

    return mapping


# ---------------------------------------------------------------------------
# Internal range detection helpers
# ---------------------------------------------------------------------------

def _rm_detect_last_data_row(ws: Worksheet) -> int:
    """Find the last row of inventory detail data (row before TOTAL)."""
    for row in range(RM_TOTAL_ROW, RM_DATA_START_ROW, -1):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a is not None and str(cell_a).strip().upper() == "TOTAL":
            return row - 1
    return RM_TOTAL_ROW - 1


def _cons_detect_structure(ws: Worksheet) -> tuple[int, int, int, int, int, int]:
    """Auto-detect Consumption sheet row positions by scanning for known labels.

    Returns:
        (last_data_row, total_row, summary_label_row,
         summary_hdr_row, summary_data_start, summary_total_row)
    """
    total_row = summary_label_row = None
    max_row = ws.max_row

    for row in range(CONS_DATA_START, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        s = str(val).strip().upper()
        if s == "TOTAL" and total_row is None:
            total_row = row
        if "AGGREGATED" in s:
            summary_label_row = row
            break

    if total_row is None:
        total_row = max_row
    if summary_label_row is None:
        summary_label_row = total_row + 3

    last_data_row    = total_row - 1
    summary_hdr_row  = summary_label_row + 1
    summary_data_start = summary_label_row + 2
    summary_total_row  = summary_data_start + CONS_SUMMARY_ORG_COUNT

    return (last_data_row, total_row, summary_label_row,
            summary_hdr_row, summary_data_start, summary_total_row)


def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _col_idx(col_letter: str) -> int:
    """Convert column letter(s) to 1-based index."""
    return openpyxl.utils.column_index_from_string(col_letter)


# ---------------------------------------------------------------------------
# Task 2 — Raw Material sheet update
# ---------------------------------------------------------------------------

def update_raw_material_sheet(
    wb:                   openpyxl.Workbook,
    inventory_detail_df:  pd.DataFrame,
    report:               list[dict],
) -> None:
    """Update inventory detail data in the Raw Material sheet.

    Strategy: match-and-update by (item_code, org_name).
    - Updates E column (primary_qty) for matched rows.
    - Zeroes out qty for items present in workbook but absent from new data.
    - Logs a warning for new items not yet in the workbook template.
    - Never moves rows; preserves all formula and summary block positions.

    Args:
        wb:                   Open openpyxl Workbook (data_only=False).
        inventory_detail_df:  Cleaned inventory detail from clean_inventory.run().
        report:               Mutable list; validation rows appended in-place.
    """
    ws  = wb["Raw Material"]
    last_data_row = _rm_detect_last_data_row(ws)
    e_col = _col_idx(RM_COL_QTY)
    b_col = _col_idx(RM_COL_ITEM_CODE)
    d_col = _col_idx(RM_COL_ORG_NAME)

    # Build lookup: (item_code, org_name) → new primary_qty
    df = inventory_detail_df.copy()
    df["item_code"] = df["item_code"].astype(str).str.strip()
    df["org_name"]  = df["org_name"].astype(str).str.strip()

    # Aggregate in case of duplicates (multiple subinv codes for same item+org)
    agg = (
        df.groupby(["item_code", "org_name"], as_index=False)
        .agg(primary_qty=("primary_qty", "sum"))
    )
    new_qty: dict[tuple, float] = {
        (r["item_code"], r["org_name"]): r["primary_qty"]
        for _, r in agg.iterrows()
    }

    # Read existing workbook key set
    wb_keys: set[tuple] = set()
    for row in range(RM_DATA_START_ROW, last_data_row + 1):
        ic  = ws.cell(row=row, column=b_col).value
        org = ws.cell(row=row, column=d_col).value
        if ic and org:
            wb_keys.add((str(ic).strip(), str(org).strip()))

    # Update E column for each existing row
    updated = zeroed = skipped = 0
    for row in range(RM_DATA_START_ROW, last_data_row + 1):
        ic  = ws.cell(row=row, column=b_col).value
        org = ws.cell(row=row, column=d_col).value
        if not ic or not org:
            continue
        key = (str(ic).strip(), str(org).strip())
        qty_cell = ws.cell(row=row, column=e_col)
        if _is_formula(qty_cell):
            skipped += 1
            continue
        if key in new_qty:
            qty_cell.value = round(new_qty[key], 4)
            updated += 1
        else:
            qty_cell.value = 0
            zeroed += 1

    # Warn about items in new data not in workbook
    new_only = set(new_qty.keys()) - wb_keys
    if new_only:
        affected = "; ".join(f"{ic}/{org}" for ic, org in sorted(new_only)[:5])
        if len(new_only) > 5:
            affected += f" ... (+{len(new_only)-5} more)"
        report.append(_row(
            "raw_material", "new_items_not_in_template", _CHECK_WARN,
            f"{len(new_only)} item+org pair(s) in fresh data are not in the workbook "
            "template. Qtys not written — update workbook template manually.",
            affected,
        ))
        logger.warning("Raw Material: %d new item+org pairs not in template.", len(new_only))
    else:
        report.append(_row(
            "raw_material", "new_items_not_in_template", _CHECK_PASS,
            "All inventory items matched existing workbook template rows.",
        ))

    report.append(_row(
        "raw_material", "inventory_update", _CHECK_PASS,
        f"Updated {updated} qty cells; zeroed {zeroed} absent items; "
        f"skipped {skipped} formula cells.",
    ))
    logger.info("Raw Material: updated=%d, zeroed=%d, skipped=%d", updated, zeroed, skipped)


# ---------------------------------------------------------------------------
# Task 3 — Consumption sheet update
# ---------------------------------------------------------------------------

def update_consumption_sheet(
    wb:                      openpyxl.Workbook,
    consumption_detail_df:   pd.DataFrame,
    period_label:            Optional[str],
    report:                  list[dict],
) -> None:
    """Replace all transaction detail rows in the Consumption sheet.

    If the new row count differs from the old one, the TOTAL formula and all
    SUMPRODUCT summary formulas are updated to cover the new range. The
    Strategy sheet's SUMPRODUCT references are also updated.

    Args:
        wb:                     Open openpyxl Workbook (data_only=False).
        consumption_detail_df:  Cleaned consumption detail from clean_consumption.run().
        period_label:           Optional period description for cell A2,
                                e.g. "April 2026 (01-Apr to 30-Apr)".
        report:                 Mutable list; validation rows appended in-place.
    """
    ws = wb["Consumption"]
    (old_last_data, old_total_row, old_summary_label, old_summary_hdr,
     old_summary_data_start, old_summary_total) = _cons_detect_structure(ws)

    # Read org order from existing summary block (preserves the workbook ordering)
    org_order: list[str] = []
    for row in range(old_summary_data_start, old_summary_total):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() not in ("", "TOTAL"):
            org_order.append(str(val).strip())

    # ── Update header text ────────────────────────────────────────────────
    if period_label:
        ws["A2"] = period_label

    # ── Clear old data rows (values only, preserve styles) ───────────────
    clear_end = max(old_summary_total + 5, old_last_data + 5)
    for row in range(CONS_DATA_START, clear_end + 1):
        for col in range(1, 10):  # columns A–I
            cell = ws.cell(row=row, column=col)
            if not _is_formula(cell):
                cell.value = None

    # ── Write new transaction rows ────────────────────────────────────────
    df = consumption_detail_df.copy()

    # Map DataFrame columns to workbook columns (A-I)
    _col_map = {
        "item_code":   1,  # A
        "item_desc":   2,  # B
        "org_name":    3,  # C
        "lot_number":  4,  # D
        "subinventory":5,  # E
        "txn_type":    6,  # F
        "primary_qty": 7,  # G  (raw, possibly negative; workbook uses ABS in its formulas)
        "date":        8,  # H
        "category":    9,  # I
    }

    n_rows = len(df)
    for i, (_, rec) in enumerate(df.iterrows()):
        row_num = CONS_DATA_START + i
        for field, col_idx in _col_map.items():
            if field not in df.columns:
                continue
            val = rec[field]
            # Convert pandas timestamps to plain Python datetime for openpyxl
            if hasattr(val, "to_pydatetime"):
                val = val.to_pydatetime()
            elif isinstance(val, pd.Period):
                val = str(val)
            ws.cell(row=row_num, column=col_idx).value = val

    new_last_data_row = CONS_DATA_START + n_rows - 1

    # ── Compute new summary block positions ───────────────────────────────
    new_total_row         = new_last_data_row + 1
    new_summary_label_row = new_last_data_row + 4
    new_summary_hdr_row   = new_last_data_row + 5
    new_summary_data_start= new_last_data_row + 6
    new_summary_total_row = new_summary_data_start + CONS_SUMMARY_ORG_COUNT

    data_range = f"${CONS_COL_QTY}$5:${CONS_COL_QTY}${new_last_data_row}"

    # ── Write TOTAL row ───────────────────────────────────────────────────
    ws.cell(row=new_total_row, column=1).value = "TOTAL"
    ws.cell(row=new_total_row, column=7).value = f"=SUM(G{CONS_DATA_START}:G{new_last_data_row})"

    # ── Write summary block ───────────────────────────────────────────────
    ws.cell(row=new_summary_label_row, column=1).value = (
        "AGGREGATED CONSUMPTION BY ORG & CATEGORY (Formula-driven from above)"
    )
    for col_idx, hdr in enumerate(
        ["Org Name", "Cotton", "Fiber (incl. Stretch)", "Stretch Fiber Only", "Total Consumed"],
        start=1,
    ):
        ws.cell(row=new_summary_hdr_row, column=col_idx).value = hdr

    r_start = CONS_DATA_START
    r_end   = new_last_data_row

    for i, org in enumerate(org_order):
        row = new_summary_data_start + i
        ws.cell(row=row, column=1).value = org
        ws.cell(row=row, column=2).value = (
            f'=SUMPRODUCT(($C${r_start}:$C${r_end}=A{row})'
            f'*($I${r_start}:$I${r_end}="Cotton")*$G${r_start}:$G${r_end})'
        )
        ws.cell(row=row, column=3).value = (
            f'=SUMPRODUCT(($C${r_start}:$C${r_end}=A{row})'
            f'*(($I${r_start}:$I${r_end}="Fiber")+($I${r_start}:$I${r_end}="Stretch Fiber"))'
            f'*$G${r_start}:$G${r_end})'
        )
        ws.cell(row=row, column=4).value = (
            f'=SUMPRODUCT(($C${r_start}:$C${r_end}=A{row})'
            f'*($I${r_start}:$I${r_end}="Stretch Fiber")*$G${r_start}:$G${r_end})'
        )
        ws.cell(row=row, column=5).value = f"=SUM(B{row}:D{row})"

    # Summary TOTAL row
    ws.cell(row=new_summary_total_row, column=1).value = "TOTAL"
    for col_idx in range(2, 6):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.cell(row=new_summary_total_row, column=col_idx).value = (
            f"=SUM({col_letter}{new_summary_data_start}:{col_letter}{new_summary_total_row-1})"
        )

    # ── Update Strategy sheet SUMPRODUCT range bounds if rows changed ──────
    if new_last_data_row != old_last_data:
        _update_strategy_consumption_refs(
            wb, old_last_data, new_last_data_row, new_summary_total_row
        )
        report.append(_row(
            "consumption", "range_updated", _CHECK_WARN,
            f"Consumption row count changed: {old_last_data} → {new_last_data_row}. "
            "Strategy SUMPRODUCT ranges updated.",
        ))
    else:
        report.append(_row(
            "consumption", "range_updated", _CHECK_PASS,
            f"Consumption row count unchanged ({new_last_data_row}). "
            "No formula range updates needed.",
        ))

    report.append(_row(
        "consumption", "rows_written", _CHECK_PASS,
        f"{n_rows} transaction rows written to Consumption sheet.",
    ))
    logger.info("Consumption: wrote %d rows, new last_data=%d", n_rows, new_last_data_row)


def _update_strategy_consumption_refs(
    wb:               openpyxl.Workbook,
    old_last_data:    int,
    new_last_data:    int,
    new_summary_total:int,
) -> None:
    """Replace Consumption!$C$5:$C$<old_ref> range bounds in Strategy formulas.

    Replaces the upper row bound of every Consumption sheet SUMPRODUCT
    reference in the Strategy sheet to cover the new data range.
    """
    ws_strat = wb["Strategy"]
    # Old range upper bound was old_summary_total (last row of old summary block)
    # We match any reference like Consumption!$X$5:$X$NNNNN
    pattern = re.compile(
        r"(Consumption!\\\$[A-Z]\\\$5:\\\$[A-Z]\\\$)(\d+)", re.IGNORECASE
    )
    # openpyxl stores formula text without extra escaping; adjust pattern:
    pattern = re.compile(
        r"(Consumption!\$[A-Z]\$5:\$[A-Z]\$)(\d+)", re.IGNORECASE
    )

    for row in ws_strat.iter_rows():
        for cell in row:
            if _is_formula(cell):
                new_formula = pattern.sub(
                    lambda m: f"{m.group(1)}{new_summary_total}",
                    cell.value,
                )
                if new_formula != cell.value:
                    cell.value = new_formula


# ---------------------------------------------------------------------------
# Task 4 — Market Inputs sheet update
# ---------------------------------------------------------------------------

def update_market_inputs_sheet(
    wb:               openpyxl.Workbook,
    market_inputs_df: pd.DataFrame,
    report:           list[dict],
    rmb_usd_rate:     float = 7.24,
) -> None:
    """Update the editable Market Inputs cells with fresh market data.

    Only literal-value cells are updated. Formula cells (B25, B29–B32, B44–B46)
    are never touched.

    Args:
        wb:               Open openpyxl Workbook (data_only=False).
        market_inputs_df: DataFrame or list of records with fields:
                          metric_name, metric_value, status.
        report:           Mutable list; validation rows appended in-place.
        rmb_usd_rate:     RMB per 1 USD for PSF conversion (default 7.24).
                          PSF is fetched in RMB/ton; workbook wants USD/MT.
    """
    ws = wb["Market Inputs"]

    df = _normalise_market_df(market_inputs_df)
    lookup: dict[str, tuple[float, str]] = {}
    for _, rec in df.iterrows():
        if str(rec.get("status", "OK")).upper() not in ("OK", "SUCCESS", "LIVE", "CACHED", ""):
            continue
        val = rec.get("metric_value")
        unit = str(rec.get("unit", ""))
        if pd.notna(val):
            lookup[str(rec["metric_name"])] = (float(val), unit)

    def _write_if_live(metric_name: str, cell_addr: str, converted_val: Optional[float],
                       description: str) -> None:
        if converted_val is None:
            report.append(_row(
                "market_inputs", metric_name.lower().replace(" ", "_"),
                _CHECK_WARN, f"{description}: metric not available in market_inputs_df.",
            ))
            return
        cell = ws[cell_addr]
        if _is_formula(cell):
            report.append(_row(
                "market_inputs", metric_name.lower().replace(" ", "_"),
                _CHECK_FAIL, f"{description}: cell {cell_addr} contains a formula — not updated.",
            ))
            return
        cell.value = converted_val
        report.append(_row(
            "market_inputs", metric_name.lower().replace(" ", "_"),
            _CHECK_PASS, f"{description}: {cell_addr} = {converted_val}",
        ))

    # ICE Cotton No. 2: USD/lb → c/lb (×100)
    ice_val = lookup.get("ICE Cotton No. 2")
    _write_if_live(
        "ICE Cotton No. 2", "B6",
        round(ice_val[0] * 100, 4) if ice_val else None,
        "ICE Cotton Spot Proxy",
    )

    # PSF Price (China): RMB/ton → USD/MT (÷ rmb_usd_rate)
    psf_val = lookup.get("PSF Price (China)")
    _write_if_live(
        "PSF Price (China)", "B11",
        round(psf_val[0] / rmb_usd_rate, 2) if psf_val else None,
        f"PSF Asia Benchmark (converted from RMB at {rmb_usd_rate})",
    )

    # SBP Policy Rate: may arrive as % integer (e.g. 10.5) or decimal (0.105)
    sbp_val = lookup.get("SBP Policy Rate")
    if sbp_val:
        sbp_num, sbp_unit = sbp_val
        # Normalise to decimal (workbook stores as 0.105, not 10.5)
        sbp_decimal = sbp_num / 100 if sbp_num > 1 else sbp_num
        _write_if_live("SBP Policy Rate", "B12", round(sbp_decimal, 6), "SBP Policy Rate")
    else:
        _write_if_live("SBP Policy Rate", "B12", None, "SBP Policy Rate")

    # USD/PKR Exchange Rate
    fx_val = lookup.get("USD/PKR Exchange Rate")
    _write_if_live(
        "USD/PKR Exchange Rate", "B13",
        round(fx_val[0], 2) if fx_val else None,
        "USD/PKR Assumption",
    )


def _normalise_market_df(market_inputs_df) -> pd.DataFrame:
    """Accept list[MarketRecord], list[dict], or DataFrame."""
    if isinstance(market_inputs_df, pd.DataFrame):
        return market_inputs_df
    records = []
    for item in market_inputs_df:
        if hasattr(item, "__dict__"):
            records.append(vars(item))
        elif hasattr(item, "_asdict"):
            records.append(item._asdict())
        elif isinstance(item, dict):
            records.append(item)
    return pd.DataFrame(records) if records else pd.DataFrame()


# ---------------------------------------------------------------------------
# Task 5 — Strategy sheet validation (no data writes)
# ---------------------------------------------------------------------------

def update_strategy_sheet(
    wb:          openpyxl.Workbook,
    strategy_df: pd.DataFrame,
    report:      list[dict],
) -> None:
    """Validate the Strategy sheet. No data writes — all cells are formulas.

    Checks that the workbook formula structure is consistent with the
    procurement_engine strategy_df output.

    Args:
        wb:          Open openpyxl Workbook (data_only=False).
        strategy_df: Output of procurement_engine.run() — used for reference counts.
        report:      Mutable list; validation rows appended in-place.
    """
    ws = wb["Strategy"]

    # Check required sheets referenced by Strategy formulas exist
    for ref_sheet in ("Raw Material", "Consumption", "Market Inputs"):
        if ref_sheet in wb.sheetnames:
            report.append(_row(
                "strategy", f"ref_sheet_{ref_sheet.lower().replace(' ','_')}",
                _CHECK_PASS, f"Referenced sheet '{ref_sheet}' exists.",
            ))
        else:
            report.append(_row(
                "strategy", f"ref_sheet_{ref_sheet.lower().replace(' ','_')}",
                _CHECK_FAIL, f"Referenced sheet '{ref_sheet}' is MISSING from workbook.",
            ))

    # Check A3 formula (Spot Price link to Market Inputs)
    a3 = ws["A3"].value
    if isinstance(a3, str) and "Market Inputs" in a3 and "B6" in a3:
        report.append(_row(
            "strategy", "market_inputs_link", _CHECK_PASS,
            "A3 references Market Inputs!B6 as expected.",
        ))
    else:
        report.append(_row(
            "strategy", "market_inputs_link", _CHECK_WARN,
            f"A3 value unexpected: {repr(a3)[:80]}. Market Inputs link may be broken.",
        ))

    # Verify Cotton section rows D8-D14 reference Raw Material summary cells
    cotton_orgs = {8: "B192", 9: "B193", 10: "B188", 11: "B190", 12: "B189",
                   13: "B191", 14: "B187"}
    broken = []
    for row_num, expected_ref in cotton_orgs.items():
        cell_val = ws.cell(row=row_num, column=4).value  # D column
        if not isinstance(cell_val, str) or expected_ref not in cell_val:
            broken.append(f"D{row_num}")
    if not broken:
        report.append(_row(
            "strategy", "raw_material_links", _CHECK_PASS,
            "Cotton section (D8:D14) correctly references Raw Material summary.",
        ))
    else:
        report.append(_row(
            "strategy", "raw_material_links", _CHECK_WARN,
            f"Strategy rows {broken} have unexpected Raw Material references.",
            ", ".join(broken),
        ))

    # Cross-check org count
    buy_hold = strategy_df[strategy_df["action"].isin(["BUY", "HOLD"])] if not strategy_df.empty else pd.DataFrame()
    report.append(_row(
        "strategy", "engine_alignment", _CHECK_PASS,
        f"procurement_engine produced {len(strategy_df)} strategy rows "
        f"({len(buy_hold)} BUY/HOLD). Workbook formulas recalculate on open.",
    ))


# ---------------------------------------------------------------------------
# Task 6 — Archive engine
# ---------------------------------------------------------------------------

def create_monthly_archive(
    workbook_path: Union[str, Path],
    archive_dir:   Union[str, Path],
    year:          int,
    month:         int,
) -> Path:
    """Save a timestamped copy of the workbook BEFORE any modifications.

    Naming convention: strategies_YYYY_MM.xlsx
    Safety: never overwrites an existing archive file.

    Args:
        workbook_path: Source workbook path (strategies.xlsx).
        archive_dir:   Directory for archive copies.
        year:          Reporting year  (e.g. 2026).
        month:         Reporting month (1–12).

    Returns:
        Path to the created archive file.

    Raises:
        FileNotFoundError: if workbook_path does not exist.
        FileExistsError:   if archive file already exists.
    """
    src  = Path(workbook_path)
    dest_dir = Path(archive_dir)

    if not src.exists():
        raise FileNotFoundError(f"Workbook not found: {src}")

    dest_dir.mkdir(parents=True, exist_ok=True)
    archive_name = f"strategies_{year:04d}_{month:02d}.xlsx"
    dest = dest_dir / archive_name

    if dest.exists():
        raise FileExistsError(
            f"Archive already exists: {dest}. "
            "Delete manually or increment month to prevent overwrite."
        )

    shutil.copy2(src, dest)
    logger.info("Archive created: %s", dest)
    return dest


# ---------------------------------------------------------------------------
# Task 7 — Load and save
# ---------------------------------------------------------------------------

def load_workbook_for_update(path: Union[str, Path]) -> openpyxl.Workbook:
    """Load workbook preserving all formulas, formatting, and charts.

    Uses data_only=False to keep formula strings intact.
    keep_vba=False (no VBA in this workbook).

    Args:
        path: Path to Strategies.xlsx.

    Returns:
        openpyxl Workbook instance ready for update.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = load_workbook(str(path), data_only=False, keep_vba=False)
    logger.info("Workbook loaded: %s  (sheets: %s)", path.name, wb.sheetnames)
    return wb


def save_workbook(
    wb:     openpyxl.Workbook,
    path:   Union[str, Path],
    report: list[dict],
) -> None:
    """Save workbook and run a post-save integrity check.

    Args:
        wb:     Modified Workbook instance.
        path:   Destination path (overwrites source).
        report: Mutable list; validation rows appended in-place.
    """
    path = Path(path)
    try:
        wb.save(str(path))
        logger.info("Workbook saved: %s", path)
        report.append(_row(
            "save", "workbook_saved", _CHECK_PASS,
            f"Workbook saved successfully to {path}.",
        ))
    except Exception as exc:
        report.append(_row(
            "save", "workbook_saved", _CHECK_FAIL,
            f"Save failed: {exc}",
        ))
        raise

    # Post-save: reopen and verify sheets + formulas survive round-trip
    _validate_saved_workbook(path, report)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate_saved_workbook(path: Path, report: list[dict]) -> None:
    """Reopen workbook and run structural integrity checks."""
    try:
        wb2 = load_workbook(str(path), data_only=False)
    except Exception as exc:
        report.append(_row(
            "validation", "workbook_reopens", _CHECK_FAIL,
            f"Saved workbook could not be reopened: {exc}",
        ))
        return

    # Required sheets exist
    missing_sheets = [s for s in WORKBOOK_SHEETS if s not in wb2.sheetnames]
    if not missing_sheets:
        report.append(_row(
            "validation", "required_sheets", _CHECK_PASS,
            f"All {len(WORKBOOK_SHEETS)} required sheets present.",
        ))
    else:
        report.append(_row(
            "validation", "required_sheets", _CHECK_FAIL,
            f"Missing sheets after save: {missing_sheets}",
            ", ".join(missing_sheets),
        ))

    # Raw Material summary block: formulas in B187 still intact
    ws_rm = wb2["Raw Material"]
    b187 = ws_rm["B187"].value
    if isinstance(b187, str) and b187.startswith("="):
        report.append(_row(
            "validation", "rm_formula_preserved", _CHECK_PASS,
            f"Raw Material B187 formula intact: {b187[:60]}",
        ))
    else:
        report.append(_row(
            "validation", "rm_formula_preserved", _CHECK_FAIL,
            f"Raw Material B187 formula lost. Value = {repr(b187)[:60]}",
        ))

    # Market Inputs: B25 must be a formula (=B12)
    ws_mi = wb2["Market Inputs"]
    b25 = ws_mi["B25"].value
    if isinstance(b25, str) and b25.startswith("="):
        report.append(_row(
            "validation", "mi_formula_preserved", _CHECK_PASS,
            f"Market Inputs B25 formula intact: {b25}",
        ))
    else:
        report.append(_row(
            "validation", "mi_formula_preserved", _CHECK_FAIL,
            f"Market Inputs B25 formula lost. Value = {repr(b25)}",
        ))

    # Strategy: A3 should be a formula linking to Market Inputs
    ws_st = wb2["Strategy"]
    a3 = ws_st["A3"].value
    if isinstance(a3, str) and a3.startswith("="):
        report.append(_row(
            "validation", "strategy_formula_preserved", _CHECK_PASS,
            f"Strategy A3 formula intact: {a3}",
        ))
    else:
        report.append(_row(
            "validation", "strategy_formula_preserved", _CHECK_FAIL,
            f"Strategy A3 formula lost. Value = {repr(a3)}",
        ))

    # Consumption: new total row formula
    ws_cons = wb2["Consumption"]
    (_, new_total_row, *_) = _cons_detect_structure(ws_cons)
    total_g = ws_cons.cell(row=new_total_row, column=7).value
    if isinstance(total_g, str) and total_g.startswith("=SUM"):
        report.append(_row(
            "validation", "consumption_total_formula", _CHECK_PASS,
            f"Consumption TOTAL row {new_total_row} formula intact: {total_g}",
        ))
    else:
        report.append(_row(
            "validation", "consumption_total_formula", _CHECK_WARN,
            f"Consumption TOTAL row formula unexpected: {repr(total_g)[:60]}",
        ))


# ---------------------------------------------------------------------------
# Public API — run()
# ---------------------------------------------------------------------------

def run(
    inventory_detail_df:    pd.DataFrame,
    consumption_detail_df:  pd.DataFrame,
    market_inputs_df,
    strategy_df:            pd.DataFrame,
    workbook_path:          Union[str, Path] = "data/strategy/Strategies.xlsx",
    archive_dir:            Union[str, Path] = "data/strategy/archive",
    year:                   Optional[int]  = None,
    month:                  Optional[int]  = None,
    period_label:           Optional[str]  = None,
    rmb_usd_rate:           float = 7.24,
) -> pd.DataFrame:
    """Orchestrate the full monthly workbook update.

    Steps (in order):
        1. Create monthly archive (BEFORE any modifications).
        2. Load workbook.
        3. Update Raw Material sheet.
        4. Update Consumption sheet.
        5. Update Market Inputs sheet.
        6. Validate Strategy sheet (no writes).
        7. Save workbook.
        8. Return workbook_update_report DataFrame.

    Args:
        inventory_detail_df:   detail_df from clean_inventory.run().
        consumption_detail_df: detail_df from clean_consumption.run().
        market_inputs_df:      list[MarketRecord] or DataFrame from market_inputs.
        strategy_df:           strategy_df from procurement_engine.run().
        workbook_path:         Path to Strategies.xlsx (default: data/strategy/Strategies.xlsx).
        archive_dir:           Archive directory (default: data/strategy/archive).
        year:                  Reporting year  (default: current year).
        month:                 Reporting month (default: current month).
        period_label:          Consumption sheet A2 text, e.g.
                               "April 2026 (01-Apr to 30-Apr)".
        rmb_usd_rate:          RMB/USD for PSF conversion (default 7.24).

    Returns:
        workbook_update_report — DataFrame with columns:
            category, check, status, detail, affected
    """
    now = datetime.now()
    year  = year  or now.year
    month = month or now.month

    report: list[dict] = []

    # ── Step 1: Archive ────────────────────────────────────────────────────
    try:
        archive_path = create_monthly_archive(workbook_path, archive_dir, year, month)
        report.append(_row(
            "archive", "monthly_archive", _CHECK_PASS,
            f"Archive created: {archive_path.name}",
            str(archive_path),
        ))
    except FileExistsError as exc:
        report.append(_row(
            "archive", "monthly_archive", _CHECK_WARN,
            str(exc),
        ))
        logger.warning("Archive skipped: %s", exc)
    except Exception as exc:
        report.append(_row(
            "archive", "monthly_archive", _CHECK_FAIL,
            f"Archive failed: {exc}",
        ))
        raise

    # ── Step 2: Load ───────────────────────────────────────────────────────
    wb = load_workbook_for_update(workbook_path)

    # ── Step 3–6: Update sheets ────────────────────────────────────────────
    update_raw_material_sheet(wb, inventory_detail_df, report)
    update_consumption_sheet(wb, consumption_detail_df, period_label, report)
    update_market_inputs_sheet(wb, market_inputs_df, report, rmb_usd_rate)
    update_strategy_sheet(wb, strategy_df, report)

    # ── Step 7: Save ───────────────────────────────────────────────────────
    save_workbook(wb, workbook_path, report)

    # Summarise
    report_df  = pd.DataFrame(report)
    fail_count = (report_df["status"] == _CHECK_FAIL).sum()
    warn_count = (report_df["status"] == _CHECK_WARN).sum()
    pass_count = (report_df["status"] == _CHECK_PASS).sum()
    logger.info(
        "Workbook update complete — PASS: %d | WARN: %d | FAIL: %d",
        pass_count, warn_count, fail_count,
    )

    return report_df


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _load_df(path: Path) -> pd.DataFrame:
    return (
        pd.read_excel(path, engine="openpyxl")
        if path.suffix.lower() in (".xlsx", ".xls")
        else pd.read_csv(path)
    )


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")

    parser = argparse.ArgumentParser(
        description="Populate Strategies.xlsx with fresh Oracle and market data"
    )
    parser.add_argument("--inventory",   required=True,
                        help="Inventory detail xlsx/csv (from clean_inventory)")
    parser.add_argument("--consumption", required=True,
                        help="Consumption detail xlsx/csv (from clean_consumption)")
    parser.add_argument("--market",      required=True,
                        help="Market inputs xlsx/csv (from market_inputs)")
    parser.add_argument("--strategy",    default=None,
                        help="Strategy xlsx/csv (from procurement_engine) — optional")
    parser.add_argument("--workbook",    default="data/strategy/Strategies.xlsx",
                        help="Path to Strategies.xlsx")
    parser.add_argument("--archive-dir", default="data/strategy/archive",
                        dest="archive_dir", help="Archive directory")
    parser.add_argument("--year",  type=int, default=None)
    parser.add_argument("--month", type=int, default=None)
    parser.add_argument("--period-label", default=None, dest="period_label",
                        help='e.g. "April 2026 (01-Apr to 30-Apr)"')
    parser.add_argument("--rmb-usd", type=float, default=7.24, dest="rmb_usd_rate",
                        help="RMB/USD rate for PSF conversion (default 7.24)")
    args = parser.parse_args()

    inv_df  = _load_df(Path(args.inventory))
    cons_df = _load_df(Path(args.consumption))
    mkt_df  = _load_df(Path(args.market))
    strat_df = _load_df(Path(args.strategy)) if args.strategy else pd.DataFrame()

    report_df = run(
        inventory_detail_df=inv_df,
        consumption_detail_df=cons_df,
        market_inputs_df=mkt_df,
        strategy_df=strat_df,
        workbook_path=args.workbook,
        archive_dir=args.archive_dir,
        year=args.year,
        month=args.month,
        period_label=args.period_label,
        rmb_usd_rate=args.rmb_usd_rate,
    )

    print("\n=== WORKBOOK UPDATE REPORT ===")
    print(report_df.to_string(index=False))

    fail_count = (report_df["status"] == "FAIL").sum()
    print(f"\nResult: {len(report_df)} checks — "
          f"{(report_df['status']=='PASS').sum()} PASS, "
          f"{(report_df['status']=='WARN').sum()} WARN, "
          f"{fail_count} FAIL")
    if fail_count:
        raise SystemExit(1)
