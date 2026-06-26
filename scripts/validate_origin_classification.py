"""
validate_origin_classification.py
-----------------------------------
Diagnostic utility for origin_classifier.py.

Runs the LOCAL/IMPORTED/UNKNOWN classifier against a real inventory
detail_df and reports coverage. Intended to be rerun every time a fresh
Oracle export is cleaned, to confirm classification coverage hasn't
degraded (e.g. a new supplier/country description format the classifier
does not yet recognise).

Does NOT modify clean_inventory.py, clean_consumption.py, or
run_forecasts.py. Reads only from clean_inventory.py's public run()
function and origin_classifier.py's public API.

Two input modes:
    --input      Path to a raw Oracle export (MG_STOCK_TILL_DATE_MULTIPLE_UNIT.xlsx).
                 Routed through clean_inventory.run() to produce detail_df.
    --workbook   Path to the strategy workbook (Strategies.xlsx). Reads the
                 Raw Material sheet directly -- it is populated from the
                 same detail_df shape by update_strategy_workbook.py, so it
                 doubles as a live snapshot of current production inventory
                 without requiring a fresh Oracle pull.

Usage:
    python validate_origin_classification.py --input "D:/.../MG_STOCK_TILL_DATE_MULTIPLE_UNIT.xlsx"
    python validate_origin_classification.py --workbook "data/strategy/Strategies.xlsx"
    python validate_origin_classification.py --workbook "..." --overrides scripts/origin_overrides.csv
    python validate_origin_classification.py --workbook "..." --report-out reports/origin_classification_report.md

Outputs (printed always; written to --report-out when given):
    - total cotton tons, local tons, imported tons, unknown tons
    - classification coverage %  (target: >= 95%)
    - unmatched (UNKNOWN) descriptions, sorted by tonnage descending
    - per-org coverage breakdown
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))

from origin_classifier import (
    IMPORTED,
    IN_SCOPE_CATEGORY,
    LOCAL,
    UNKNOWN,
    aggregate_by_origin,
    classify_dataframe,
    load_overrides_csv,
)

COVERAGE_TARGET_PCT = 95.0

# Raw Material sheet layout -- matches update_strategy_workbook.py's
# RM_COL_* constants and RM_DATA_START_ROW / RM_TOTAL_ROW.
_RM_DATA_START_ROW = 6
_RM_TOTAL_ROW_LABEL = "TOTAL"
_RM_COL_ITEM_CODE = 2   # B
_RM_COL_DESC = 3        # C
_RM_COL_ORG_NAME = 4    # D
_RM_COL_QTY = 5         # E
_RM_COL_CATEGORY = 6    # F


# ---------------------------------------------------------------------------
# Input loaders
# ---------------------------------------------------------------------------

def load_detail_from_oracle_export(
    input_path: str | Path,
    mapping_csv: Optional[str | Path] = None,
) -> pd.DataFrame:
    """Load detail_df via the existing clean_inventory.py pipeline (unmodified)."""
    from clean_inventory import run as clean_inventory_run

    detail_df, _summary_df = clean_inventory_run(input_path, mapping_csv=mapping_csv)
    return detail_df


def load_detail_from_workbook(workbook_path: str | Path) -> pd.DataFrame:
    """Reconstruct a detail_df-shaped DataFrame from the Strategy workbook's
    Raw Material sheet.

    The Raw Material sheet is populated from inventory_detail_df by
    update_strategy_workbook.py using the exact same columns (item_code,
    description, org_name, primary_qty, category) -- see RM_COL_* constants
    in that module. Reading it back gives a live snapshot of current
    production inventory without needing a fresh Oracle export, and without
    touching clean_inventory.py.
    """
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    ws = wb["Raw Material"]

    last_data_row = None
    for row in range(_RM_DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip().upper() == _RM_TOTAL_ROW_LABEL:
            last_data_row = row - 1
            break
    if last_data_row is None:
        last_data_row = ws.max_row

    records = []
    for row in range(_RM_DATA_START_ROW, last_data_row + 1):
        item_code = ws.cell(row=row, column=_RM_COL_ITEM_CODE).value
        description = ws.cell(row=row, column=_RM_COL_DESC).value
        org_name = ws.cell(row=row, column=_RM_COL_ORG_NAME).value
        qty = ws.cell(row=row, column=_RM_COL_QTY).value
        category = ws.cell(row=row, column=_RM_COL_CATEGORY).value
        if item_code is None or category is None:
            continue
        records.append({
            "item_code": str(item_code).strip(),
            "description": str(description).strip() if description is not None else "",
            "org_name": str(org_name).strip() if org_name is not None else "",
            "primary_qty": float(qty) if qty is not None else 0.0,
            "subinventory_code": "",
            "category": str(category).strip(),
        })

    return pd.DataFrame.from_records(records)


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def build_coverage_report(classified_df: pd.DataFrame) -> dict:
    """Compute the full coverage report from a classified detail_df.

    Returns a dict with keys:
        total_tons, local_tons, imported_tons, unknown_tons,
        coverage_pct, meets_target, unmatched_descriptions (DataFrame),
        per_org_coverage (DataFrame)
    """
    cotton = classified_df[classified_df["category"] == IN_SCOPE_CATEGORY].copy()

    total_tons = float(cotton["primary_qty"].sum())
    local_tons = float(cotton.loc[cotton["origin"] == LOCAL, "primary_qty"].sum())
    imported_tons = float(cotton.loc[cotton["origin"] == IMPORTED, "primary_qty"].sum())
    unknown_tons = float(cotton.loc[cotton["origin"] == UNKNOWN, "primary_qty"].sum())

    coverage_pct = ((local_tons + imported_tons) / total_tons * 100) if total_tons else 0.0

    unmatched = (
        cotton[cotton["origin"] == UNKNOWN]
        .groupby(["item_code", "description"], as_index=False)
        .agg(tons=("primary_qty", "sum"))
        .sort_values("tons", ascending=False)
        .reset_index(drop=True)
    )

    per_org = (
        cotton
        .assign(is_classified=lambda d: d["origin"].isin([LOCAL, IMPORTED]))
        .groupby("org_name")
        .apply(lambda g: pd.Series({
            "total_tons": g["primary_qty"].sum(),
            "local_tons": g.loc[g["origin"] == LOCAL, "primary_qty"].sum(),
            "imported_tons": g.loc[g["origin"] == IMPORTED, "primary_qty"].sum(),
            "unknown_tons": g.loc[g["origin"] == UNKNOWN, "primary_qty"].sum(),
            "coverage_pct": (
                g.loc[g["is_classified"], "primary_qty"].sum() / g["primary_qty"].sum() * 100
                if g["primary_qty"].sum() else 0.0
            ),
        }))
        .reset_index()
        .sort_values("coverage_pct")
    )

    return {
        "total_tons": total_tons,
        "local_tons": local_tons,
        "imported_tons": imported_tons,
        "unknown_tons": unknown_tons,
        "coverage_pct": coverage_pct,
        "meets_target": coverage_pct >= COVERAGE_TARGET_PCT,
        "unmatched_descriptions": unmatched,
        "per_org_coverage": per_org,
    }


def print_report(report: dict) -> None:
    print("\n" + "=" * 72)
    print("ORIGIN CLASSIFICATION COVERAGE REPORT")
    print("=" * 72)
    print(f"  Total Cotton inventory : {report['total_tons']:>14,.1f} tons")
    print(f"  LOCAL                  : {report['local_tons']:>14,.1f} tons "
          f"({report['local_tons']/report['total_tons']*100:5.1f}%)" if report['total_tons'] else "")
    print(f"  IMPORTED                : {report['imported_tons']:>14,.1f} tons "
          f"({report['imported_tons']/report['total_tons']*100:5.1f}%)" if report['total_tons'] else "")
    print(f"  UNKNOWN                 : {report['unknown_tons']:>14,.1f} tons "
          f"({report['unknown_tons']/report['total_tons']*100:5.1f}%)" if report['total_tons'] else "")
    print("-" * 72)
    status = "PASS" if report["meets_target"] else "FAIL"
    print(f"  Coverage: {report['coverage_pct']:.2f}%  "
          f"(target >= {COVERAGE_TARGET_PCT:.0f}%)  [{status}]")
    print("=" * 72)

    if not report["unmatched_descriptions"].empty:
        print("\nUNMATCHED DESCRIPTIONS (UNKNOWN bucket, sorted by tonnage):")
        print(report["unmatched_descriptions"].to_string(index=False))
    else:
        print("\nNo unmatched descriptions -- 100% of Cotton rows classified.")

    print("\nPER-ORG COVERAGE:")
    print(report["per_org_coverage"].to_string(index=False))


def write_markdown_report(report: dict, output_path: str | Path) -> None:
    lines = [
        "# Origin Classification Coverage Report",
        "",
        f"- Total Cotton inventory: **{report['total_tons']:,.1f} tons**",
        f"- LOCAL: {report['local_tons']:,.1f} tons "
        f"({report['local_tons']/report['total_tons']*100:.1f}%)" if report['total_tons'] else "- LOCAL: 0",
        f"- IMPORTED: {report['imported_tons']:,.1f} tons "
        f"({report['imported_tons']/report['total_tons']*100:.1f}%)" if report['total_tons'] else "- IMPORTED: 0",
        f"- UNKNOWN: {report['unknown_tons']:,.1f} tons "
        f"({report['unknown_tons']/report['total_tons']*100:.1f}%)" if report['total_tons'] else "- UNKNOWN: 0",
        "",
        f"**Coverage: {report['coverage_pct']:.2f}%** "
        f"(target >= {COVERAGE_TARGET_PCT:.0f}%) "
        f"-- {'PASS' if report['meets_target'] else 'FAIL'}",
        "",
        "## Unmatched Descriptions",
        "",
    ]
    if report["unmatched_descriptions"].empty:
        lines.append("None -- 100% of Cotton rows classified.")
    else:
        lines.append("| item_code | description | tons |")
        lines.append("|---|---|---|")
        for _, r in report["unmatched_descriptions"].iterrows():
            lines.append(f"| {r['item_code']} | {r['description']} | {r['tons']:,.1f} |")

    lines += ["", "## Per-Org Coverage", "", "| org_name | total_tons | local_tons | imported_tons | unknown_tons | coverage_pct |",
              "|---|---|---|---|---|---|"]
    for _, r in report["per_org_coverage"].iterrows():
        lines.append(
            f"| {r['org_name']} | {r['total_tons']:,.1f} | {r['local_tons']:,.1f} | "
            f"{r['imported_tons']:,.1f} | {r['unknown_tons']:,.1f} | {r['coverage_pct']:.1f}% |"
        )

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"\nReport written -> {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Validate origin_classifier.py coverage against real inventory data"
    )
    parser.add_argument("--input", default=None,
                         help="Path to raw Oracle export (MG_STOCK_TILL_DATE_MULTIPLE_UNIT.xlsx)")
    parser.add_argument("--workbook", default=None,
                         help="Path to Strategies.xlsx (reads Raw Material sheet directly)")
    parser.add_argument("--overrides", default=None,
                         help="Optional item_code->origin override CSV")
    parser.add_argument("--mapping-csv", default=None, dest="mapping_csv",
                         help="Optional commodity_mapping.csv passed through to clean_inventory")
    parser.add_argument("--report-out", default=None, dest="report_out",
                         help="Optional path to write a markdown coverage report")
    args = parser.parse_args()

    if not args.input and not args.workbook:
        parser.error("Provide either --input (raw Oracle export) or --workbook (Strategies.xlsx)")

    if args.input:
        detail_df = load_detail_from_oracle_export(args.input, args.mapping_csv)
    else:
        detail_df = load_detail_from_workbook(args.workbook)

    overrides = load_overrides_csv(args.overrides) if args.overrides else None
    classified_df = classify_dataframe(detail_df, overrides=overrides)
    report = build_coverage_report(classified_df)

    print_report(report)

    if args.report_out:
        write_markdown_report(report, args.report_out)

    sys.exit(0 if report["meets_target"] else 1)
